#!/usr/bin/env python3
"""
Generate nf_template.xlsx with all 5 sheets:
  NF MSHIP, NF S, NF LOANS, NF FS, NF FARM

Headers match the parser's constants exactly (case-insensitive).
Usage: python3 docs/scripts/generate_template.py
Output: nf_template.xlsx (in repo root)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_BORDER = Border(
    top=Side(style="thin", color="B4C6E7"),
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def style_sheet(ws, headers):
    """Write header row with styling and auto-width columns."""
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = HDR_ALIGN
        c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(i)].width = max(len(h) + 3, 14)


def main():
    wb = openpyxl.Workbook()

    # ── Sheet 1: NF MSHIP ──────────────────────────────────────────
    ws = wb.active
    ws.title = "NF MSHIP"
    headers_m = [
        "member_id", "join_date", "status", "exit_date", "gender",
        "age_group", "region", "urban_rural", "agm_attendance",
        "leadership_role", "voting_exercised",
    ]
    style_sheet(ws, headers_m)
    ws.append(["M001", "2023-01-15", "Active", "", "Male",
               "18-35", "Hhohho", "Urban", "TRUE", "Treasurer", "TRUE"])
    ws.append(["M002", "2023-03-01", "Active", "2024-12-31", "Female",
               "36-50", "Manzini", "Rural", "FALSE", "", "FALSE"])

    # ── Sheet 2: NF S ─────────────────────────────────────────────
    ws2 = wb.create_sheet("NF S")
    headers_s = [
        "member_id", "savings_account_id", "account_type",
        "account_opening_date", "account_status", "contribution_frequency",
        "last_contribution_date", "number_of_contributions", "balance_trend",
        "zero_balance_flag", "withdrawal_frequency_category",
        "emergency_withdrawals_flag", "interest_rate", "balance",
    ]
    style_sheet(ws2, headers_s)
    ws2.append(["M001", "SAV001", "Voluntary", "2023-01-15", "Active",
                "Monthly", "2025-06-01", 12, "Stable", "FALSE",
                "Low", "FALSE", "3.5", "2500.00"])
    ws2.append(["M002", "SAV002", "Mandatory", "2023-03-01", "Active",
                "Monthly", "2025-06-15", 24, "Increasing", "FALSE",
                "None", "FALSE", "4.0", "5000.00"])

    # ── Sheet 3: NF LOANS ─────────────────────────────────────────
    ws3 = wb.create_sheet("NF LOANS")
    headers_l = [
        "member_id", "loan_id", "loan_product_type", "loan_start_date",
        "loan_maturity_date", "loan_status", "borrower_type",
        "youth_borrower_flag", "women_borrower_flag", "rural_borrower_flag",
        "repayment_regularity", "days_past_due_category",
        "missed_installments_count", "restructured_loan_flag",
        "number_of_restructurings", "early_settlement_flag",
        "multiple_loans_flag", "large_borrower_flag", "interest_rate",
        "balance", "loan_amount",
    ]
    style_sheet(ws3, headers_l)
    ws3.append(["M001", "LN001", "Agricultural", "2024-01-01", "2025-12-31",
                "Performing", "Individual", "TRUE", "FALSE", "FALSE",
                "Regular", "0", 0, "FALSE", 0, "FALSE", "FALSE", "FALSE",
                "8.5", "15000.00", "20000.00"])
    ws3.append(["M002", "LN002", "SME", "2024-06-01", "2026-05-31",
                "Performing", "Group", "FALSE", "TRUE", "TRUE",
                "Regular", "0", 0, "FALSE", 0, "FALSE", "FALSE", "FALSE",
                "10.0", "8000.00", "10000.00"])

    # ── Sheet 4: NF FS ────────────────────────────────────────────
    ws4 = wb.create_sheet("NF FS")
    headers_fd = [
        "member_id", "fixed_deposit_id", "deposit_type", "start_date",
        "maturity_date", "status", "tenure_category",
        "original_tenure_selected", "early_withdrawal_flag",
        "rollover_at_maturity_flag", "number_of_renewals",
        "change_in_tenure_at_renewal", "single_depositor_dependency_flag",
        "interest_rate", "balance",
    ]
    style_sheet(ws4, headers_fd)
    ws4.append(["M001", "FD001", "Fixed Term", "2024-01-01", "2025-01-01",
                "Active", "12 Months", "12 Months", "FALSE", "TRUE",
                1, "FALSE", "FALSE", "6.0", "10000.00"])
    ws4.append(["M002", "FD002", "Fixed Term", "2025-06-01", "2026-06-01",
                "Active", "12 Months", "12 Months", "FALSE", "FALSE",
                0, "FALSE", "FALSE", "5.5", "5000.00"])

    # ── Sheet 5: NF FARM ──────────────────────────────────────────
    ws5 = wb.create_sheet("NF FARM")
    headers_fc = [
        "cooperative_type",
        "primary_activities",
        "year_of_establishment",
        "operational_status",
        "active_producer_flag",
        "production_type",
        "participation_frequency",
        "delivery_compliance",
        "production_cycle_type",
        "use_of_production_planning",
        "use_of_shared_inputs",
        "quality_compliance_flag",
        "market_channel_type",
        "formal_offtake_agreement",
        "buyer_concentration_flag",
        "price_predictability_category",
        "access_to_storage",
        "access_to_processing_facilities",
        "transport_coordination",
        "climate_exposure_type",
        "irrigation_access",
        "climate_mitigation_practices",
    ]
    style_sheet(ws5, headers_fc)

    # Row 2 — A diversified farming cooperative (all flags TRUE, established)
    ws5.append([
        "Farmer Cooperative",           # cooperative_type
        "Mixed Farming",                # primary_activities
        2010,                           # year_of_establishment
        "Active",                       # operational_status
        "TRUE",                         # active_producer_flag
        "Crop & Livestock",             # production_type
        "Monthly",                      # participation_frequency
        "Full",                         # delivery_compliance
        "Seasonal",                     # production_cycle_type
        "TRUE",                         # use_of_production_planning
        "TRUE",                         # use_of_shared_inputs
        "TRUE",                         # quality_compliance_flag
        "Local Market",                 # market_channel_type
        "TRUE",                         # formal_offtake_agreement
        "FALSE",                        # buyer_concentration_flag
        "Moderate",                     # price_predictability_category
        "TRUE",                         # access_to_storage
        "TRUE",                         # access_to_processing_facilities
        "Shared Transport",             # transport_coordination
        "Low",                          # climate_exposure_type
        "TRUE",                         # irrigation_access
        "Crop Rotation, Cover Cropping", # climate_mitigation_practices
    ])

    # Row 3 — A smaller cooperative with fewer resources
    ws5.append([
        "Farming Cooperative",          # cooperative_type
        "Vegetable Farming",            # primary_activities
        2018,                           # year_of_establishment
        "Active",                       # operational_status
        "TRUE",                         # active_producer_flag
        "Crops Only",                   # production_type
        "Quarterly",                    # participation_frequency
        "Partial",                      # delivery_compliance
        "Perennial",                    # production_cycle_type
        "FALSE",                        # use_of_production_planning
        "TRUE",                         # use_of_shared_inputs
        "FALSE",                        # quality_compliance_flag
        "Direct Sales",                 # market_channel_type
        "FALSE",                        # formal_offtake_agreement
        "TRUE",                         # buyer_concentration_flag
        "Volatile",                     # price_predictability_category
        "FALSE",                        # access_to_storage
        "FALSE",                        # access_to_processing_facilities
        "Individual Transport",         # transport_coordination
        "Medium",                       # climate_exposure_type
        "FALSE",                        # irrigation_access
        "None",                         # climate_mitigation_practices
    ])

    # Save
    out_path = os.path.join(os.path.dirname(__file__), "..", "..", "nf_template.xlsx")
    wb.save(out_path)
    print(f"✅ Template saved to {os.path.abspath(out_path)}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Rows: NF MSHIP=2, NF S=2, NF LOANS=2, NF FS=2, NF FARM=2")


if __name__ == "__main__":
    main()
